import datetime
import numbers
import os
import sqlite3

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class Dimension:
    def __init__(self, minRow, maxRow, minColumn, maxColumn):
        self.minRow = minRow
        self.maxRow = maxRow
        self.minColumn = minColumn
        self.maxColumn = maxColumn

    def __repr__(self):
        return 'minRow: {0}, maxRow: {1}, minColumn: {2}, maxColumn: {3}'.format(self.minRow,
                                                                                 self.maxRow,
                                                                                 self.minColumn,
                                                                                 self.maxColumn)

class Member:
    def __init__(self, name, hidden):
        self.id = -1
        self.name = name
        self.hidden = hidden
        self.remaining = 0

    def __repr__(self):
        return 'name: {0}, hidden: {1}'.format(self.name,
                                               self.hidden)

class Cost:
    def __init__(self, member, money):
        self.member = member
        self.money = money

    def __repr__(self):
        return 'member: {0}, money: {1}'.format(self.member,
                                                self.money)

class LunchEvent:
    def __init__(self, index, restaurant, time):
        self.id = -1
        self.index = index
        self.restaurant = restaurant
        self.time = time
        self.costs = []

    def __repr__(self):
        return 'index: {0}, restaurant: {1}, time: {2}, sum: {3}'.format(self.index,
                                                                           self.restaurant,
                                                                           self.time,
                                                                           self.sum)

    @property
    def sum(self):
        costSum = 0
        for oneCost in self.costs:
            costSum += oneCost.money
        return costSum

def loadFromXLSX(xlsxName):
    wb = load_workbook(xlsxName, data_only=True)
    sheet = wb['Lunch']

    dimension = Dimension(sheet.min_row,
                          sheet.max_row,
                          sheet.min_column,
                          sheet.max_column)

    hiddenColumns = []
    for colLetter,colDimension in sheet.column_dimensions.items():
        if colDimension.hidden:
            for index in range(colDimension.min, colDimension.max + 1):
                hiddenColumns.append(get_column_letter(index))

    members = {}
    for memberRow in sheet.iter_rows(min_row=1, max_row=1,
                                     min_col=2, max_col=(dimension.maxColumn - 1)):
        for cell in memberRow:
            isHidden = cell.column in hiddenColumns
            members[cell.column] = Member(cell.value, isHidden)

    for memberRow in sheet.iter_rows(min_row=dimension.maxRow, max_row=dimension.maxRow,
                                     min_col=2, max_col=(dimension.maxColumn - 1)):
        for cell in memberRow:
            if not isinstance(cell.value, numbers.Number):
                print 'unexpected cell [{0}, {1}], expecting number, but get {2}'.format(cell.row, cell.column, cell.value)
                return
            members[cell.column].remaining = cell.value
    print members

    lunchEvents = []

    restaurant = ''
    expectingDate = False
    for recordRow in sheet.iter_rows(min_row=2, max_row=(dimension.maxRow - 1),
                                     min_col=1, max_col=1):
        for cell in recordRow:
            if not cell.is_date:
                if expectingDate:
                    print 'unexpected event cell [{0}, {1}], expecting restaurant name'.format(cell.row, cell.column)
                    return
                restaurant = cell.value
                expectingDate = True
            else:
                if not expectingDate:
                    print 'unexpected event cell [{0}, {1}], expecting date'.format(cell.row, cell.column)
                    return
                lunchEvents.append(LunchEvent(cell.row - 1, restaurant, cell.value))
                expectingDate = False

    for theEvent in lunchEvents:
        for costCol in sheet.iter_rows(min_row=theEvent.index, max_row=theEvent.index,
                                       min_col=2, max_col=(dimension.maxColumn - 1)):
            for cell in costCol:
                if cell.value != None:
                    if not isinstance(cell.value, numbers.Number):
                        print 'unexpected cell [{0}, {1}], expecting number, but get {2}'.format(cell.row, cell.column, cell.value)
                        return
                    theEvent.costs.append(Cost(members[cell.column], cell.value))
    return members.values(), lunchEvents

def create_db(con):
    con.execute("DROP TABLE IF EXISTS lunch_foundation_event")
    con.execute("DROP TABLE IF EXISTS lunch_foundation_member")
    con.execute("DROP TABLE IF EXISTS lunch_foundation_cost")
    con.execute("CREATE TABLE IF NOT EXISTS lunch_foundation_event ( \
                                    id INTEGER PRIMARY KEY, \
                                    restaurant VARCHAR, \
                                    time DATE)")
    con.execute("CREATE TABLE IF NOT EXISTS lunch_foundation_member ( \
                                    id INTEGER PRIMARY KEY, \
                                    name VARCHAR, \
                                    hidden INTEGER, \
                                    remaining REAL)")
    con.execute("CREATE TABLE IF NOT EXISTS lunch_foundation_cost ( \
                                    event_id INTEGER NOT NULL, \
                                    member_id INTEGER NOT NULL, \
                                    cost REAL, \
                                    recharge REAL, \
                                    UNIQUE(event_id, member_id))")

def insert_members(con, members):
    for oneMember in members:
        isHiddenMember = 0
        if oneMember.hidden:
            isHiddenMember = 1

        result = con.execute("INSERT OR REPLACE INTO lunch_foundation_member ( \
                                    name, \
                                    hidden, \
                                    remaining \
                                ) VALUES (?, ?, ?)",
                                (oneMember.name, isHiddenMember, oneMember.remaining))
        oneMember.id = result.lastrowid

def insert_costs(con, events):
    c = con.cursor()
    for oneEvent in events:
        for oneCost in oneEvent.costs:
            c.execute("INSERT OR REPLACE INTO lunch_foundation_cost ( \
                                    event_id, \
                                    member_id, \
                                    cost, \
                                    recharge \
                                ) VALUES (?, ?, ?, ?)",
                                (oneEvent.id, oneCost.member.id, oneCost.money, 0))

    con.commit()

def insert_events(con, events):
    for oneEvent in events:
        result = con.execute("INSERT OR REPLACE INTO lunch_foundation_event ( \
                                    restaurant, \
                                    time \
                                ) VALUES (?, ?)",
                                (oneEvent.restaurant, oneEvent.time))
        oneEvent.id = result.lastrowid

    insert_costs(con, events)

def saveToDB(dbName, members, events):
    print 'saving to {0}'.format(dbName)
    with sqlite3.connect(dbName) as con:
        # create_db(con)

        insert_members(con, members)
        insert_events(con, events)

    print 'Done'

def main():
    members, lunchEvents = loadFromXLSX('TeamLunchFoundation.xlsx')
    # print '\n'.join(map(lambda x: str(x), lunchEvents))
    print 'total: {0} events, {1} members'.format(len(lunchEvents), len(members))

    saveToDB('TeamLunchFoundation.db', members, lunchEvents)

if __name__ == "__main__":
    main()
